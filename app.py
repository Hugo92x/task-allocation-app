import streamlit as st
import pandas as pd
import json
import re
import openpyxl
import tempfile
import os
import base64

# Function to safely serialize JSON data
def json_serialize_safe(obj):
    if pd.isna(obj):
        return None
    if isinstance(obj, (pd.Timestamp, pd.Timedelta)):
        return str(obj)
    return str(obj)

# Function to clean location data
def clean_location(location):
    if not location:
        return 'Unknown Location'
    return re.sub(r'\s*\([A-Z]\)(?:\s*,\s*[^,\]]*)*', '', location)

# Function to determine time period (dagdeel)
def determine_dagdeel(shift_time):
    try:
        start_time = shift_time.split('-')[0].replace('+1', '')
        hour = int(start_time.split(':')[0])
        minute = int(start_time.split(':')[1])
    except ValueError:
        return 'Unknown'
    
    time_in_minutes = hour * 60 + minute
    
    if 300 <= time_in_minutes <= 540:  # 05:00-09:30
        return 'Ochtend'
    elif 541 <= time_in_minutes <= 690:  # 09:31-11:30
        return 'Tussen'
    elif 691 <= time_in_minutes <= 1170:  # 11:31-19:30
        return 'Avond'
    elif 1171 <= time_in_minutes <= 1500 or time_in_minutes < 299:  # 19:31-01:00+1 or 00:00-04:59
        return 'Nacht'
    else:
        return 'Unknown'

# Function to parse shift cell data
def parse_shift_cell(cell_value):
    if pd.isna(cell_value) or str(cell_value).strip() == '':
        return None, None, None
        
    cell_value = str(cell_value).replace('_x000D_', '').strip()
    
    if not cell_value or cell_value.lower() == 'file':
        return None, None, None
    
    if cell_value.startswith('[') and cell_value.endswith(']'):
        location = clean_location(cell_value[1:-1])
        return None, None, location
    
    match = re.match(r'(\d{2}:\d{2})(?:\+1)?-(\d{2}:\d{2})(?:\+1)?\s*(?:\[(.*?)\])?', cell_value)
    
    if not match:
        return None, None, None
    
    start_time, end_time, location = match.groups()
    location = clean_location(location)
    
    return start_time, end_time, location

# Function to read employee schedule from Excel
def read_employee_schedule(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb['Medewerkers']
        
        df = pd.read_excel(file_path, sheet_name='Medewerkers', header=None)
        st.write(f"Total columns in sheet: {df.shape[1]}")
        
        date_row = df.iloc[7]
        processed_records = []
        
        excluded_colors = ['FFFF00', 'FF3B3B', '00FFFF', 'FFFFFF00', 'FFFF3B3B', 'FF00FFFF', 'FFA9D4']
        training_colors = ['33CCCC', 'FF33CCCC']
        
        for col in range(3, df.shape[1]):
            date_str = date_row[col]
            if pd.isna(date_str):
                continue
                
            date = pd.to_datetime(date_str, format='%d-%m-%Y')
            
            for row in range(8, df.shape[0]):
                if pd.isna(df.iloc[row, 0]):
                    continue
                
                first_name = str(df.iloc[row, 0])
                last_name = str(df.iloc[row, 1])
                function = str(df.iloc[row, 2])
                cell_value = df.iloc[row, col]
                
                cell_coord = sheet.cell(row=row+1, column=col+1)
                cell_fill = cell_coord.fill
                cell_color = cell_fill.start_color.rgb if cell_fill.start_color.rgb else None
                
                if not cell_color:
                    continue
                    
                if any(cell_color.endswith(color[-6:]) for color in excluded_colors):
                    continue
                    
                if pd.isna(cell_value) or cell_value == '':
                    continue
                    
                is_training = any(cell_color.endswith(color[-6:]) for color in training_colors)
                
                start_time, end_time, location = parse_shift_cell(cell_value)
                if start_time is None:
                    continue
                
                # Skip training shifts that start at 08:30 or 09:00
                if is_training and start_time in ['08:30', '09:00']:
                    continue
                    
                dagdeel = determine_dagdeel(f"{start_time}-{end_time}")
                
                record = {
                    'Medewerkers': f"{first_name} {last_name}".strip(),
                    'DefaultTask': 'Meelopen' if is_training else None,  # Changed from 'Training / Meelopen'
                    'Functie': function,
                    'Dag': date.strftime('%A'),
                    'Datum': date.strftime('%Y-%m-%d'),
                    'Starttijd': start_time,
                    'Eindtijd': end_time,
                    'Locatie': location,
                    'Dagdeel': dagdeel,
                    'CellColor': cell_color,
                    'IsTrainee': is_training  # Flag to identify trainees for UI interactions
                }
                processed_records.append(record)
        
        st.write(f"Total processed records: {len(processed_records)}")
        return pd.DataFrame(processed_records)
        
    except Exception as e:
        st.error(f"Error reading employee schedule: {str(e)}")
        raise

# Function to read daily tasks from Excel
def read_daily_tasks(file_path):
    days = ['Maandag', 'Dinsdag', 'Woensdag', 'Donderdag', 'Vrijdag', 'Zaterdag', 'Zondag']
    period_order = {'Ochtend': 1, 'Avond': 2, 'Nacht': 3}
    function_order = {'CC': 1, 'TL': 2, 'DC': 3, 'A': 4, 'B': 5, 'C': 6, 'D': 7, 'E+': 8, 'E': 9}
    all_tasks = []
    
    # Initialize task counters for unique ID generation
    task_counters = {}
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    def determine_period(start_time):
        if not start_time:
            return ''
        try:
            hour = int(start_time.split(':')[0])
            if 22 <= hour <= 23 or hour < 6:  # 22:00-05:59
                return 'Nacht'
            elif 6 <= hour < 14:  # 06:00-13:59
                return 'Ochtend'
            else:  # 14:00-21:59
                return 'Avond'
        except (ValueError, IndexError):
            return ''
    
    for day in days:
        try:
            sheet = wb[f'Taken {day}']
            df = pd.read_excel(file_path, sheet_name=f'Taken {day}')
            if not df.empty:
                start_row = 2
                
                for idx, row in df.iterrows():
                    if pd.isna(row.iloc[0]):
                        continue
                        
                    cell = sheet.cell(row=start_row + idx, column=1)
                    cell_color = cell.fill.start_color.rgb if cell.fill.start_color.rgb else None
                    
                    # Get task name from column A
                    task_desc = str(row.iloc[0]).strip()
                    if pd.isna(task_desc) or task_desc == '':
                        continue
                    
                    parts = task_desc.split('\n')
                    task_name = parts[0].strip()
                    
                    # Get function from column B
                    function = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else ''
                    
                    # Get start time from column C
                    start_time = str(row.iloc[2]) if not pd.isna(row.iloc[2]) else ''
                    if start_time:
                        start_time = ':'.join(start_time.split(':')[:2])  # Only HH:MM
                        
                    # Get end time from column D
                    end_time = str(row.iloc[3]) if not pd.isna(row.iloc[3]) else ''
                    if end_time:
                        end_time = ':'.join(end_time.split(':')[:2])  # Only HH:MM
                    
                    # Get location from column F
                    location = str(row.iloc[5]) if not pd.isna(row.iloc[5]) else ''
                    
                    # Determine period from start time
                    period = determine_period(start_time)
                    
                    # Create a unique task key for counting
                    task_key = f"{task_name}_{start_time}_{end_time}_{period}"
                    
                    # Update counter for this task key
                    if task_key not in task_counters:
                        task_counters[task_key] = 0
                    else:
                        task_counters[task_key] += 1
                    
                    # Create unique task identifier including the counter
                    task_id = f"{task_key}_{task_counters[task_key]}"
                    
                    task = {
                        'TaskName': task_name,
                        'Function': function,
                        'Time': f"{start_time} - {end_time}" if start_time and end_time else '',
                        'Locatie': location,
                        'Day': day,
                        'Dagdeel': period,
                        'CellColor': cell_color,
                        'TaskId': task_id
                    }
                    
                    if task['TaskName'] and task['TaskName'] != 'nan':
                        all_tasks.append(task)
                
        except Exception as e:
            st.warning(f"Error reading {day} tasks: {str(e)}")
            continue
    
    # Create DataFrame and sort
    df_tasks = pd.DataFrame(all_tasks) if all_tasks else pd.DataFrame()
    if not df_tasks.empty:
        # First sort by period
        df_tasks['PeriodOrder'] = df_tasks['Dagdeel'].map(period_order)
        # Then by function
        df_tasks['FunctionOrder'] = df_tasks['Function'].map(function_order)
        
        # Sort by period first, then function
        df_tasks = df_tasks.sort_values(['PeriodOrder', 'FunctionOrder'])
        
        # Remove helper columns but keep TaskId to maintain uniqueness
        df_tasks = df_tasks.drop(['PeriodOrder', 'FunctionOrder'], axis=1)
    
    return df_tasks

# Function to generate HTML content
def generate_html(employees_df, tasks_df):
    period_order = {'Ochtend': 1, 'Tussen': 2, 'Avond': 3, 'Nacht': 4}
    
    sorted_df = employees_df.sort_values(
        by=['Datum', 'Dagdeel', 'Functie'], 
        key=lambda x: x.map(period_order) if x.name == 'Dagdeel' else x
    )
    
    sorted_df = sorted_df.map(json_serialize_safe)
    tasks_df = tasks_df.map(json_serialize_safe)
    
    def convert_color(color):
        if not color or color == 'FFFFFFFF':
            return None
        return f'#{color[2:]}' if color.startswith('FF') else f'#{color}'
    
    employees_data = sorted_df.to_dict('records')
    tasks_data = tasks_df.to_dict('records') if not tasks_df.empty else []
    
    for record in employees_data:
        record['CellColor'] = convert_color(record.get('CellColor'))
    
    for task in tasks_data:
        if task.get('CellColor'):
            task['CellColor'] = convert_color(task.get('CellColor'))

    # Generate the HTML content (same as your original script)
    html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Weekly Task Allocation</title>
    <style>
        body {{ 
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 5pt;
            box-sizing: border-box;
        }}
        
        .header-row {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10pt;
        }}
        
        .page-container {{
            display: flex;
            flex-direction: column;
            width: 100%;
        }}
        
        /* Main containers */
        .day-container {{
            width: 100%;
            margin-bottom: 10pt;
        }}
        
        .day-title {{
            background-color: #f9f9f9;
            padding: 5pt;
            border-radius: 1pt;
            margin-bottom: 10pt;
        }}
        
        .period-container {{
            width: 100%;
            margin-bottom: 10pt;
        }}
        
        .content-row {{
            display: flex;
            width: 100%;
        }}
        
        .employee-section {{ 
            width: 65%;
            padding-right: 10pt;
            box-sizing: border-box;
        }}
        
        .task-section {{
            width: 35%;
            padding-left: 10pt;
            box-sizing: border-box;
        }}
        
        /* Filter and button row */
        .controls-row {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10pt;
        }}
        
        /* Filter controls */
        .filters {{ 
            display: flex;
            gap: 5pt;
            align-items: center;
            margin-bottom: 10pt;
        }}
        
        .filter-label {{
            margin-right: 5pt;
        }}
        
        /* Action buttons */
        .action-buttons {{
            display: flex;
            gap: 10pt;
        }}
        
        .action-button {{
            background-color: #5b99d2;
            color: white;
            border: none;
            border-radius: 2pt;
            padding: 4pt 8pt;
            font-weight: bold;
            cursor: pointer;
            font-size: 0.9em;
            transition: background-color 0.2s;
        }}
        
        .action-button:hover {{
            background-color: #4a7eb3;
        }}
        
            .print-button {{
        background-color: #dc3545;
    }}
    
    .print-button:hover {{
        background-color: #c82333;
    }}
    
    .export-button {{
        background-color: #007bff;
    }}
    
    .export-button:hover {{
        background-color: #0069d9;
    }}
    
    .timeline-button {{
        background-color: #6c757d;
    }}
    
    .timeline-button:hover {{
        background-color: #5a6268;
    }}
    
    .auto-allocate-button {{
        background-color: #28a745;
    }}
    
    .auto-allocate-button:hover {{
        background-color: #218838;
    }}
    
    .unassign-button {{
        background-color: #ffc107;
        color: #212529;
    }}
    
    .unassign-button:hover {{
        background-color: #e0a800;
    }}

        .spinner {{
            display: inline-block;
            width: 14px;
            height: 14px;
            border: 2px solid rgba(255,255,255,0.3);
            border-radius: 50%;
            border-top-color: white;
            animation: spin 0.8s ease-in-out infinite;
            display: none;
        }}

        @keyframes spin {{
            to {{ transform: rotate(360deg); }}
        }}
        
        /* Headers */
        .period-header {{
            padding: 4pt;
            margin: 0 0 0pt 0;
            font-weight: bold;
            color: #fff;
            border-top-left-radius: 2pt;
            border-top-right-radius: 2pt;
            display: flex;
            align-items: center;
            min-height: 20pt;
            box-sizing: border-box;
            width: 100%;
        }}
        
        .period-header.ochtend {{
            background-color: #45a1ff;
        }}
        
        .period-header.avond {{
            background-color: #40c057;
        }}
        
        .period-header.nacht {{
            background-color: #fd7e14;
        }}
        
        /* Tables */
        table {{ 
            width: 100%;
            border-collapse: collapse;
            margin: 0;
            font-size: 0.8em;
            table-layout: fixed;
            margin-bottom: 5pt;
        }}
        
        th, td {{ 
            border: 1pt solid #ddd;
            padding: 1pt 2pt;
            text-align: left;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        
        th {{
            background-color: #f8f9fa;
        }}
        
        /* Row background colors */
        .ochtend-row {{
            background-color: #E3F2FD;
        }}
        
        .avond-row {{
            background-color: #E8F5E9;
        }}
        
        .nacht-row {{
            background-color: #FFF3E0;
        }}
        
        /* Task items */
        .task-item {{
            padding: 1pt 2pt;
            margin: 1pt;
            border: 1pt solid #ccc;
            background-color: #fff;
            border-radius: 2pt;
            cursor: move;
            position: relative;
        }}
        
        .draggable-task {{
            cursor: move;
            padding: 1pt 2pt;
            margin: 1pt;
            border: 1pt solid #ccc;
            background-color: #fff;
            border-radius: 2pt;
        }}
        
        .task-dropzone {{
            min-height: 10pt;
            border: 1pt dashed #ccc;
            margin: 1pt;
            border-radius: 1pt;
            padding: 1pt;
        }}
        
        .task-dropzone.dragover {{
            background-color: #e1f5fe;
            border-color: #2196f3;
        }}
        
        .task-return-button {{
            cursor: pointer;
            color: #ff4444;
            position: absolute;
            top: 1pt;
            right: 2pt;
            font-weight: bold;
            padding: 0 2pt;
            line-height: 1;
            z-index: 10;
        }}
        
        .task-return-button:hover {{
            color: #cc0000;
        }}
        
        .task-details {{
            font-size: 0.9em;
            color: #666;
            margin-top: 2pt;
        }}
        
        .task-cell-light {{
            background-color: rgba(255, 255, 255, 0.7);
        }}
        
        .task-cell-details {{
            font-size: 0.8em;
            color: #666;
            margin-top: 2pt;
        }}
        
        /* Headers */
        h1 {{
            margin-top: 0;
            margin-bottom: 0;
        }}
        
        h3 {{
            margin-top: 0;
            margin-bottom: 10pt;
        }}
        
        /* Badge for counts */
        .count-badge {{
            font-size: 0.85em;
            font-weight: normal;
            margin-left: 5pt;
        }}
        
        /* Function matching alerts */
        .function-incompatible-alert {{
            font-family: Arial, sans-serif;
        }}

        .alert-content {{
            display: flex;
            flex-direction: column;
            align-items: center;
        }}

        .alert-content strong {{
            font-size: 1.2em;
            margin-bottom: 10pt;
            color: #d9534f;
        }}

        .alert-content p {{
            margin: 10pt 0;
            text-align: center;
        }}

        .alert-close-button {{
            margin-top: 10pt;
            padding: 5pt 15pt;
            background-color: #5bc0de;
            color: white;
            border: none;
            border-radius: 4pt;
            cursor: pointer;
            font-weight: bold;
        }}

        .alert-close-button:hover {{
            background-color: #31b0d5;
        }}

        /* Visual indicator for incompatible tasks when dragging */
        .task-incompatible {{
            opacity: 0.6;
            box-shadow: 0 0 5px #ff4444;
        }}

        /* Visual effect for dropzone when incompatible task is dragged over */
        .dropzone-incompatible {{
            background-color: rgba(255, 200, 200, 0.3) !important;
            border-color: #ff4444 !important;
        }}
        
        /* For dragging effect */
        .dragging {{
            opacity: 0.5;
        }}

        /* Task name styling */
        .task-name {{
            padding-right: 20pt; /* Make room for the X button */
            display: block;
        }}
        
        /* For printing */
        @media print {{
            .action-buttons, .filters, button {{
                display: none !important;
            }}
            
            body {{
                padding: 0;
                margin: 0;
            }}
            
            .day-title {{
                page-break-before: always;
            }}
            
            .period-container {{
                break-inside: avoid;
            }}
        }}
    </style>
</head>
<body>
<!-- Updated header row with title and reordered action buttons -->
<div class="header-row">
    <h1>Weekly Task Allocation</h1>
    <div class="action-buttons">
        <button id="printPlanningButton" class="action-button print-button">
            Print Planning
        </button>
        <button id="exportFlightScheduleButton" class="action-button export-button">
            Export Flight Schedule
        </button>
        <button id="timelineViewButton" class="action-button timeline-button">
            Timeline-view
        </button>
        <button id="autoAllocateButton" class="action-button auto-allocate-button">
            <span id="allocateSpinner" class="spinner"></span>
            Auto-Allocate Tasks
        </button>
        <button id="unassignAllButton" class="action-button unassign-button">
            Unassign Tasks
        </button>
    </div>
</div>
    
    <div class="filters">
        <span class="filter-label">Date:</span>
        <select id="dateFilter">
            <option value="all">All Dates</option>
        </select>
        
        <span class="filter-label">Location:</span>
        <select id="locationFilter">
            <option value="all">All Locations</option>
        </select>
        
        <span class="filter-label">Period:</span>
        <select id="periodFilter">
            <option value="all">All Periods</option>
            <option value="Ochtend">Ochtend</option>
            <option value="Tussen">Tussen</option>
            <option value="Avond">Avond</option>
            <option value="Nacht">Nacht</option>
        </select>
    </div>

    <div class="page-container" id="pageContainer"></div>

<script>
        (function() {{
            const scheduleData = {json.dumps(employees_data)};
            const tasksData = {json.dumps(tasks_data)};
            const taskAssignmentsByEmployee = new Map();
            let isUpdating = false;

            document.addEventListener('DOMContentLoaded', function() {{
                console.log('DOM loaded and ready');
                const uniqueDates = [...new Set(scheduleData.map(entry => entry.Datum))].sort();
                const uniqueLocations = [...new Set(scheduleData.map(entry => entry.Locatie))].sort();

                const dateFilter = document.getElementById('dateFilter');
                const locationFilter = document.getElementById('locationFilter');
                const periodFilter = document.getElementById('periodFilter');

                uniqueDates.forEach(date => {{
                    const option = document.createElement('option');
                    option.value = date;
                    option.textContent = date;
                    dateFilter.appendChild(option);
                }});

                uniqueLocations.forEach(location => {{
                    const option = document.createElement('option');
                    option.value = location;
                    option.textContent = location;
                    locationFilter.appendChild(option);
                }});

                if (uniqueDates.length > 0) {{
                    dateFilter.value = uniqueDates[0];
                }}

                dateFilter.addEventListener('change', updateDisplay);
                locationFilter.addEventListener('change', updateDisplay);
                periodFilter.addEventListener('change', updateDisplay);
                
                // Add event listeners for all action buttons
                document.getElementById('autoAllocateButton').addEventListener('click', autoAllocateTasks);
                
                document.getElementById('exportFlightScheduleButton').addEventListener('click', function() {{
                    alert('Export Flight Schedule functionality will be implemented here');
                    // TODO: Implement export functionality
                }});
                
                document.getElementById('timelineViewButton').addEventListener('click', showTimelineView);
                
                document.getElementById('printPlanningButton').addEventListener('click', function() {{
    createPrintModal();
                }});
                
                // Add event listener for the unassign all button
                document.getElementById('unassignAllButton').addEventListener('click', function() {{
                    // Get the current date, location, and period filter values
                    const selectedDate = document.getElementById('dateFilter').value;
                    const selectedLocation = document.getElementById('locationFilter').value;
                    const selectedPeriod = document.getElementById('periodFilter').value;
                    
                    if (selectedDate === 'all') {{
                        alert('Please select a specific date first');
                        return;
                    }}
                    
                    // Confirm the user wants to unassign all tasks
                    if (confirm('Are you sure you want to unassign all tasks for the current selection?')) {{
                        // Get all employees matching the filters
                        let filteredEmployees = [...scheduleData];
                        
                        // Filter by date
                        filteredEmployees = filteredEmployees.filter(function(entry) {{
                            return entry.Datum === selectedDate;
                        }});
                        
                        // Apply location filter if specified
                        if (selectedLocation !== 'all') {{
                            filteredEmployees = filteredEmployees.filter(function(entry) {{
                                return entry.Locatie === selectedLocation;
                            }});
                        }}
                        
                        // Apply period filter if specified
                        if (selectedPeriod !== 'all') {{
                            filteredEmployees = filteredEmployees.filter(function(entry) {{
                                return entry.Dagdeel === selectedPeriod;
                            }});
                        }}
                        
                        // Clear all task assignments for the filtered employees
                        for (const employee of filteredEmployees) {{
                            const key = `${{employee.Medewerkers}}-${{employee.Datum}}`;
                            if (taskAssignmentsByEmployee.has(key)) {{
                                taskAssignmentsByEmployee.delete(key);
                            }}
                        }}
                        
                        // Update the display
                        updateDisplay();
                        
                        // Show confirmation
                        alert('All tasks have been unassigned for the current selection.');
                    }}
                }});
                
                // Set up global event delegation for task return buttons
                document.addEventListener('click', function(e) {{
                    // Check if the clicked element is a return button
                    if (e.target.classList.contains('task-return-button')) {{
                        e.preventDefault();
                        e.stopPropagation();
                        
                        const taskId = e.target.getAttribute('data-task-id');
                        const date = e.target.getAttribute('data-date');
                        const employeeId = e.target.getAttribute('data-employee');
                        
                        console.log(`Unassigning task via delegation: ${{taskId}} from ${{employeeId}} on ${{date}}`);
                        unassignTask(date, taskId, employeeId);
                    }}
                }});
                
                updateDisplay();
            }});

            // Rest of the JavaScript functions from your original script go here
            // Function to determine if an employee can perform a task based on function matching
            function canEmployeePerformTask(employeeFunction, taskFunction) {{
                // Parse employee function to extract the function number
                const empFuncMatch = employeeFunction.match(/^(\d+)\./);
                if (!empFuncMatch) return false;
                
                // Extract the function number (1, 2, 3, etc.)
                const functionNumber = parseInt(empFuncMatch[1], 10);
                
                // Define the function codes based on employee function number
                const employeeFunctionCodes = {{
                    1: "CC",   // 1. Crew Chief -> CC
                    2: "TL",   // 2. Teamleader -> TL
                    3: "DC",   // 3. Deur Coördinator -> DC
                    4: "A",    // 4. WH Agent A -> A
                    5: "B",    // 5. WH Agent B -> B
                    6: "C",    // 6. WH Agent C -> C
                    7: "D",    // 7. WH Agent D -> D
                    8: "E+",   // 8. WH Agent E+ -> E+
                    9: "E"     // 9. WH Agent E -> E
                }};
                
                // Get the employee's function code (CC, TL, DC, etc.)
                const empFuncCode = employeeFunctionCodes[functionNumber] || "";
                
                // Define the function hierarchy based on your matrix
                // Each function can perform certain tasks (their own and those below them)
                const functionCapabilities = {{
                    "CC": ["CC", "TL", "DC", "A", "B", "C", "D", "E+", "E"], // 1. Crew Chief can do all
                    "TL": ["TL", "DC", "A", "B", "C", "D", "E+", "E"],      // 2. Teamleader can do all except CC
                    "DC": ["DC", "A", "B", "C", "D", "E+", "E"],            // 3. Deur Coördinator can do all except CC, TL
                    "A":  ["A", "B", "C", "D", "E+", "E"],                  // 4. WH Agent A choices
                    "B":  ["B", "C", "D", "E+", "E"],                       // 5. WH Agent B choices
                    "C":  ["C", "D", "E+", "E"],                            // 6. WH Agent C choices
                    "D":  ["D", "E+", "E"],                                 // 7. WH Agent D choices
                    "E+": ["E+", "E"],                                      // 8. WH Agent E+ choices
                    "E":  ["E"]                                             // 9. WH Agent E choices
                }};
                
                // Check if employee function exists in hierarchy
                if (!functionCapabilities[empFuncCode]) {{
                    console.error("Unknown employee function code:", empFuncCode, "from", employeeFunction);
                    return false;
                }}
                
                // Check if task function is in the list of functions the employee can perform
                const canPerform = functionCapabilities[empFuncCode].includes(taskFunction);
                
                console.log(`Employee with function ${{employeeFunction}} (${{empFuncCode}}) ${{canPerform ? 'CAN' : 'CANNOT'}} perform task requiring ${{taskFunction}} function`);
                
                return canPerform;
            }}

            // Auto-allocate tasks function
            function autoAllocateTasks() {{
    // Get the button and spinner
    const button = document.getElementById('autoAllocateButton');
    const spinner = document.getElementById('allocateSpinner');
    
    // Disable button and show spinner
    button.disabled = true;
    spinner.style.display = 'inline-block';
    
    // Get the current date, location, and period filter values
    const selectedDate = document.getElementById('dateFilter').value;
    const selectedLocation = document.getElementById('locationFilter').value;
    const selectedPeriod = document.getElementById('periodFilter').value;
    
    if (selectedDate === 'all') {{
        alert('Please select a specific date first');
        button.disabled = false;
        spinner.style.display = 'none';
        return;
    }}
    
    // Find all available tasks
    let availableTasks = tasksData.filter(function(task) {{ 
        const dayOfWeek = new Date(selectedDate).toLocaleString('nl-NL', {{weekday: 'long'}});
        return task.Day.toLowerCase() === dayOfWeek.toLowerCase();
    }});
    
    // Apply location filter if specified
    if (selectedLocation !== 'all') {{
        availableTasks = availableTasks.filter(function(task) {{
            return task.Locatie && task.Locatie.toLowerCase() === selectedLocation.toLowerCase();
        }});
    }}
    
    // Apply period filter if specified
    if (selectedPeriod !== 'all') {{
        availableTasks = availableTasks.filter(function(task) {{
            return task.Dagdeel === selectedPeriod;
        }});
    }}
    
    // Find all employees for the selected date
    let availableEmployees = scheduleData.filter(function(employee) {{ 
        return employee.Datum === selectedDate;
    }});
    
    // Apply location filter if specified
    if (selectedLocation !== 'all') {{
        availableEmployees = availableEmployees.filter(function(employee) {{
            return employee.Locatie === selectedLocation;
        }});
    }}
    
    // Apply period filter if specified
    if (selectedPeriod !== 'all') {{
        availableEmployees = availableEmployees.filter(function(employee) {{
            return employee.Dagdeel === selectedPeriod;
        }});
    }}
    
    // First, clear all existing task assignments for the filtered employees
    clearTaskAssignments(selectedDate, availableEmployees);
    
    // Sort tasks by function hierarchy (CC, TL, DC, A, B, C, D, E+, E)
    // This ensures higher-level tasks are allocated first
    const functionOrder = {{
        'CC': 1, 'TL': 2, 'DC': 3, 'A': 4, 'B': 5, 'C': 6, 'D': 7, 'E+': 8, 'E': 9
    }};
    
    availableTasks.sort(function(a, b) {{
        const funcA = functionOrder[a.Function] || 99;
        const funcB = functionOrder[b.Function] || 99;
        return funcA - funcB;
    }});
    
    // Sort employees by function hierarchy (1. Crew Chief, 2. Teamleader, etc.)
    // This ensures we try to assign tasks to higher-level employees first
    availableEmployees.sort(function(a, b) {{
        const empAMatch = a.Functie.match(/^(\d+)\./);
        const empBMatch = b.Functie.match(/^(\d+)\./);
        
        const empALevel = empAMatch ? parseInt(empAMatch[1], 10) : 99;
        const empBLevel = empBMatch ? parseInt(empBMatch[1], 10) : 99;
        
        return empALevel - empBLevel;
    }});
    
    // For storing temporary assignments
    const tempAssignments = new Map();
    const assignedTaskIds = new Set();
    const assignedEmployees = new Set(); // Keep track of employees who already have a task
    
    console.log(`Starting task allocation with ${{availableTasks.length}} tasks and ${{availableEmployees.length}} employees`);
    
    // First pass: Assign exactly one task per employee based on dagdeel match and function compatibility
    for (const task of availableTasks) {{
        if (assignedTaskIds.has(task.TaskId)) continue; // Skip already assigned tasks
        
        for (const employee of availableEmployees) {{
            // Skip employees who already have a task assigned in this first pass
            if (assignedEmployees.has(employee.Medewerkers)) continue;
            
            // Match periods (dagdeel) between employee and task
            if (employee.Dagdeel !== task.Dagdeel) {{
                continue; // Skip if periods don't match
            }}
            
            // Check if this employee can perform this task
            if (canEmployeePerformTask(employee.Functie, task.Function)) {{
                console.log(`Assigning task "${{task.TaskName}}" (${{task.Time}}) to employee ${{employee.Medewerkers}} - first task assignment`);
                
                // Create the key for this employee
                const key = `${{employee.Medewerkers}}-${{employee.Datum}}`;
                
                // Get or create the task array for this employee
                if (!tempAssignments.has(key)) {{
                    tempAssignments.set(key, []);
                }}
                
                // Assign this task to this employee
                tempAssignments.get(key).push({{
                    taskId: task.TaskId,
                    taskData: task
                }});
                
                assignedTaskIds.add(task.TaskId);
                assignedEmployees.add(employee.Medewerkers); // Mark this employee as having a task
                break; // Move to next task
            }}
        }}
    }}
    
    console.log(`First pass completed. Initial assignments made.`);
    
    // Second pass: Consider time conflicts for additional task assignments
    for (const task of availableTasks) {{
        // Skip already assigned tasks
        if (assignedTaskIds.has(task.TaskId)) continue;
        
        for (const employee of availableEmployees) {{
            // Match periods (dagdeel) between employee and task
            if (employee.Dagdeel !== task.Dagdeel) {{
                continue; // Skip if periods don't match
            }}
            
            // Check if this employee can perform this task
            if (!canEmployeePerformTask(employee.Functie, task.Function)) {{
                continue; // Skip if function incompatible
            }}
            
            // Check for time conflicts with existing tasks
            const key = `${{employee.Medewerkers}}-${{employee.Datum}}`;
            const existingTasks = tempAssignments.get(key) || [];
            
            if (hasTimeConflict(existingTasks, task)) {{
                console.log(`Time conflict detected in second pass for task "${{task.TaskName}}" (${{task.Time}}) - skipping employee ${{employee.Medewerkers}}`);
                continue; // Skip if there's a time conflict
            }}
            
            console.log(`Assigning additional task "${{task.TaskName}}" (${{task.Time}}) to employee ${{employee.Medewerkers}}`);
            
            // Get or create the task array for this employee
            if (!tempAssignments.has(key)) {{
                tempAssignments.set(key, []);
            }}
            
            // Assign this task to this employee
            tempAssignments.get(key).push({{
                taskId: task.TaskId,
                taskData: task
            }});
            
            assignedTaskIds.add(task.TaskId);
            break; // Move to next task
        }}
    }}
    
    console.log(`Second pass completed. Final assignments ready.`);
    
    // Now apply the temporary assignments to the actual data structure
    // Clear existing assignments and add all from the temp map
    taskAssignmentsByEmployee.clear();
    
    for (const [key, tasks] of tempAssignments.entries()) {{
        taskAssignmentsByEmployee.set(key, tasks);
    }}
    
    // Update the display
    setTimeout(function() {{
        updateDisplay();
        button.disabled = false;
        spinner.style.display = 'none';
        
        // Show success message
        const totalAssigned = Array.from(taskAssignmentsByEmployee.values()).reduce((total, tasks) => total + tasks.length, 0);
        alert(`Auto-allocation complete! ${{totalAssigned}} tasks have been assigned.`);
    }}, 500);
}}

            // Remaining JavaScript functions...
        }})();
    </script>
</body>
</html>
    """
    
    return html_content

# Improve the download link function to make it more prominent
def get_download_link(html_content, filename="task_allocation.html"):
    """Generates a link to download the HTML file"""
    b64 = base64.b64encode(html_content.encode()).decode()
    href = f'<a href="data:text/html;base64,{b64}" download="{filename}" class="download-button">Download Task Allocation Interface</a>'
    
    # Add some CSS to make the download button more prominent
    button_css = """
    <style>
        .download-button {
            display: inline-block;
            padding: 12px 20px;
            background-color: #4CAF50;
            color: white;
            text-align: center;
            text-decoration: none;
            font-size: 16px;
            border-radius: 4px;
            cursor: pointer;
            margin: 10px 0;
            font-weight: bold;
        }
        .download-button:hover {
            background-color: #45a049;
        }
    </style>
    """
    
    return button_css + href

# Streamlit App
st.set_page_config(page_title="Task Allocation App", layout="wide")

st.title("Task Allocation Web App")

# Add clearer instructions
st.write("""
## Instructions:
1. Upload your Excel file below
2. After processing, download the HTML file
3. Open the downloaded file in your browser for the full interactive experience
""")

# File uploader
uploaded_file = st.file_uploader("Upload Excel file", type=['xlsx'])

if uploaded_file is not None:
    st.success("File uploaded successfully!")
    
    # Save the uploaded file to a temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name
    
    try:
        with st.spinner("Processing file..."):
            # Process the file
            employees_df = read_employee_schedule(tmp_path)
            tasks_df = read_daily_tasks(tmp_path)
            
            # Generate HTML
            html_content = generate_html(employees_df, tasks_df)
            
            # Show some basic stats
            col1, col2 = st.columns(2)
            with col1:
                st.subheader("Employee Statistics")
                st.write(f"Total employees: {len(employees_df)}")
                st.write(f"Unique dates: {employees_df['Datum'].nunique()}")
                st.write(f"Locations: {', '.join(employees_df['Locatie'].unique())}")
                
            with col2:
                st.subheader("Task Statistics")
                st.write(f"Total tasks: {len(tasks_df)}")
                tasks_by_day = tasks_df['Day'].value_counts().to_dict()
                st.write("Tasks by day:", tasks_by_day)
            
            # Add a prominent download section
            st.subheader("📥 Download Your Interactive Task Allocation Interface")
            st.markdown("""
            **For the best experience with full interactivity:**
            1. Click the download button below
            2. Open the downloaded HTML file in your browser
            3. Enjoy all interactive features including drag-and-drop, auto-allocation, and timeline views
            """)
            
            # Provide download link
            st.markdown(get_download_link(html_content), unsafe_allow_html=True)
            
            # Also show a preview (optional)
            with st.expander("Show Preview (Limited Interactivity)"):
                st.warning("Note: This preview has limited interactive functionality. For the full experience, download the HTML file.")
                st.components.v1.html(html_content, height=600, scrolling=True)
            
    except Exception as e:
        st.error(f"Error processing the file: {str(e)}")
    
    # Clean up the temp file
    os.unlink(tmp_path)
else:
    st.info("Please upload an Excel file with the correct format.")
    
    # Example format information
    with st.expander("Expected Excel File Format"):
        st.write("""
        Your Excel file should have:
        
        1. A sheet named 'Medewerkers' with employee schedules
        2. Sheets named 'Taken Maandag', 'Taken Dinsdag', etc. for daily tasks
        
        The format should match the one used in your existing script.
        """)